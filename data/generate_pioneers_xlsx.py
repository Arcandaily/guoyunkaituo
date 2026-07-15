#!/usr/bin/env python3
"""
国运开拓 - 开拓者数据面板追踪Excel生成器
生成专业多sheet XLSX，用于追踪每位开拓者资源、技能、开拓币、蓝图等复杂数据变化。
运行: python generate_pioneers_xlsx.py
输出: pioneers_data_tracking.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

import datetime

wb = Workbook()

# Styles
def get_header_fill():
    return PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

def get_header_font():
    return Font(bold=True, color="FFFFFF", size=11)

def get_input_fill():
    return PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")  # Light blue for inputs

def get_formula_fill():
    return PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Light green for formulas

def get_special_fill():
    return PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Yellow for protagonist/special

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============ Sheet 1: Pioneers_Overview ============ 
ws1 = wb.active
ws1.title = "Pioneers_Overview"

# Headers
headers = ["ID", "姓名/代号", "角色", "国家", "领导力", "策略", "战斗", "科技", "修真融合度", 
           "金属", "晶体", "重氢", "电子货币", "开拓币", "当前舰船数", "蓝图数量", "成就点", "当前星系进度", "备注"]

for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=header)
    cell.font = get_header_font()
    cell.fill = get_header_fill()
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border

# Sample 11 Chinese Pioneers data (placeholder, customize per novel)
pioneers_data = [
    [1, "华国开拓者-01", "总指挥/舰队领袖", "华国", 95, 90, 75, 70, 60, 12500, 8200, 4500, 15000, 3200, 8, 12, 2450, "新手星系-阶段3", "官方选拔核心，团结力强"],
    [2, "华国开拓者-02", "后勤/资源管理", "华国", 80, 85, 60, 95, 40, 9800, 15000, 6200, 8000, 1800, 5, 8, 1200, "新手星系-阶段3", "擅长资源优化"],
    [3, "华国开拓者-03", "战斗/护航专家", "华国", 70, 75, 95, 65, 50, 4500, 3200, 2800, 5000, 950, 12, 6, 890, "新手星系-阶段3", "高光战斗记录"],
    [4, "华国开拓者-04", "情报/侦察", "华国", 85, 80, 70, 85, 55, 6200, 4100, 3500, 11000, 2100, 6, 9, 1560, "新手星系-阶段3", "情报网络构建中"],
    [5, "华国开拓者-05", "科技/工程", "华国", 75, 90, 65, 98, 45, 7800, 9500, 5100, 9500, 1400, 7, 15, 980, "新手星系-阶段3", "工程舰维护专家"],
    [6, "华国开拓者-06", "医疗/生化", "华国", 65, 70, 80, 90, 60, 3200, 2800, 1900, 4000, 750, 4, 5, 620, "新手星系-阶段3", "生化制剂研发"],
    [7, "华国开拓者-07", "外交/联盟协调", "华国", 90, 88, 55, 75, 50, 5100, 3800, 2400, 13000, 1650, 3, 7, 1100, "新手星系-阶段3", "跨国协调"],
    [8, "华国开拓者-08", "采矿/勘探", "华国", 60, 65, 85, 80, 40, 15200, 6800, 8900, 6000, 1100, 9, 4, 780, "新手星系-阶段3", "资源采集高产"],
    [9, "华国开拓者-09", "战机/护航艇操作", "华国", 70, 72, 92, 68, 55, 3800, 2500, 2100, 4500, 820, 15, 6, 950, "新手星系-阶段3", "机动战高光"],
    [10, "华国开拓者-10", "后备/多面手", "华国", 78, 82, 78, 82, 48, 6900, 5200, 3100, 8500, 1350, 6, 8, 1050, "新手星系-阶段3", "灵活支援"],
    [11, "第11开拓者-主角", "修真融合专家/特殊顾问", "华国", 88, 95, 70, 85, 98, 8500, 7200, 4800, 12500, 4500, 10, 18, 1890, "新手星系-阶段3", "双重穿越者，金手指：炼药/炼器与科技融合（低灵气环境独特加成）"],
]

for row_idx, row_data in enumerate(pioneers_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        if col_idx == 11:  # Xianxia_Fusion column - special for protagonist
            if row_idx == 12:  # Protagonist row
                cell.fill = get_special_fill()
            else:
                cell.fill = get_input_fill()
        elif col_idx in [10,11,12,13,14]:  # Resource columns - input
            cell.fill = get_input_fill()
        else:
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# Add totals row with formulas
ws1.cell(row=13, column=1, value="TOTAL").font = Font(bold=True)
for col in range(10, 18):  # Resources and counts
    col_letter = get_column_letter(col)
    cell = ws1.cell(row=13, column=col, value=f"=SUM({col_letter}2:{col_letter}12)")
    cell.font = Font(bold=True)
    cell.fill = get_formula_fill()
    cell.border = thin_border

# Column widths
for col in range(1, len(headers)+1):
    ws1.column_dimensions[get_column_letter(col)].width = 18 if col < 9 else 14

ws1.column_dimensions['B'].width = 22
ws1.column_dimensions['R'].width = 25

# ============ Sheet 2: Resource_Log (追踪变化) ============ 
ws2 = wb.create_sheet("Resource_Log")
log_headers = ["时间/卷次", "开拓者ID", "姓名", "操作/事件", "金属Δ", "晶体Δ", "重氢Δ", "电子货币Δ", "开拓币Δ", "蓝图变化", "备注"]
for col, header in enumerate(log_headers, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.font = get_header_font()
    cell.fill = get_header_fill()
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border

# Sample log entries (expand per volume)
sample_logs = [
    ["Vol1-新手星系-阶段1", 11, "第11开拓者-主角", "完成直播任务-资源勘探演示", 0, 0, 0, 0, 850, "+1低级蓝图", "人联观众打赏+任务完成"],
    ["Vol1-新手星系-阶段2", 1, "华国开拓者-01", "团队采矿行动", 3200, 1800, 950, 0, 120, "-", "团队资源共享"],
    ["Vol1-新手星系-阶段3", 11, "第11开拓者-主角", "炼器融合实验-小型无人机优化",  -150, 0, 0, 0, 0, "+2效率提升", "金手指应用：阵法优化采矿效率"],
]

for row_idx, log in enumerate(sample_logs, 2):
    for col_idx, value in enumerate(log, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        if col_idx in [5,6,7,8,9]:
            cell.fill = get_input_fill()

for col in range(1, len(log_headers)+1):
    ws2.column_dimensions[get_column_letter(col)].width = 16
ws2.column_dimensions['D'].width = 30
ws2.column_dimensions['J'].width = 20

# ============ Sheet 3: Notes_Instructions ============ 
ws3 = wb.create_sheet("使用说明")
ws3['A1'] = "国运开拓者数据面板使用说明"
ws3['A1'].font = Font(bold=True, size=14)
ws3.merge_cells('A1:F1')

instructions = [
    "1. Pioneers_Overview: 当前所有开拓者状态总览。修真融合度列（黄色高亮）为主角专属，记录金手指应用效果。",
    "2. Resource_Log: 每卷/每阶段资源与开拓币变化日志。开拓币来源：万族战场人联高级观众直播打赏、完成指定任务、成就解锁。",
    "3. 资源说明：金属、晶体、重氢 - 参考无尽的拉格朗日游戏，采矿无人机采集，用于建造/维修舰船。电子货币 - 游戏内交易货币。开拓币 - 小说原创，用于开箱抽蓝图盲盒或直接购买技术。",
    "4. 更新方式：每卷结束后更新Overview当前值，添加Log条目。使用Excel筛选/透视表分析团队资源。",
    "5. 蓝图与开箱：蓝图数量列追踪已拥有。可扩展单独Sheet记录具体蓝图ID与来源（直播/考核奖励/合成）。",
    "6. 自定义：添加新列或新开拓者行。主角金手指融合示例见备注列或单独hybrid_tech.md。",
    "7. 数据来源：舰船/资源基础数据来自无尽的拉格朗日Wiki (CC BY-NC-SA 3.0)。小说特定系统（开拓币、直播）为原创设计。",
]

for i, line in enumerate(instructions, 3):
    ws3.cell(row=i, column=1, value=line)
    ws3.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)

ws3.column_dimensions['A'].width = 120

wb.save('/home/workdir/artifacts/pioneers_data_tracking.xlsx')
print("Excel generated: pioneers_data_tracking.xlsx")