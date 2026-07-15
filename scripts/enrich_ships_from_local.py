#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""离线增强舰船表：合并 fleet 样本 + index + detail 中已有字段。"""
from __future__ import annotations

import csv
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
WIKI = ROOT / "wiki_data"
DATA = ROOT / "data"


def load_csv(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with path.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def parse_numbers_from_text(text: str) -> dict:
    if not text:
        return {}
    kv = {}
    patterns = {
        "structure": [
            r"结构值[=：:]\s*([0-9]+)",
            r"满结构[=：:]\s*([0-9]+)",
            r"Structure[_\s]?Value[=：:]\s*([0-9]+)",
            r"([0-9]{4,6})\s*\(实战",
        ],
        "armor": [r"装甲[=：:]\s*([0-9.%]+)", r"满抵抗[=：:]\s*([0-9.%]+)", r"抵抗[=：:]\s*([0-9.%]+)"],
        "shield": [r"护盾[=：:]\s*([0-9.%]+)", r"满护盾[=：:]\s*([0-9.%]+)"],
        "cruise_speed": [r"巡航[=：:]?\s*速度[=：:]\s*([0-9]+)", r"满巡航[=：:]\s*([0-9]+)", r"巡航速度\s*([0-9]+)"],
        "warp_speed": [r"曲率[=：:]?\s*速度[=：:]\s*([0-9]+)", r"满曲率[=：:]\s*([0-9]+)", r"曲率速度\s*([0-9]+)"],
        "metal": [r"满金属[=：:]\s*([0-9]+)", r"金属[=：:]\s*([0-9]+)", r"Build_Metal[=：:]\s*([0-9]+)"],
        "crystal": [r"满晶体[=：:]\s*([0-9]+)", r"晶体[=：:]\s*([0-9]+)"],
        "deuterium": [r"满重氢[=：:]\s*([0-9]+)", r"重氢[=：:]\s*([0-9]+)"],
        "ship_dps": [r"对舰[=：:]\s*([0-9]+)", r"Firepower_vs_Ship[=：:]\s*([0-9]+)"],
        "air_dps": [r"对空[=：:]\s*([0-9]+)", r"Firepower_vs_Air[=：:]\s*([0-9]+)"],
        "siege_dps": [r"攻城[=：:]\s*([0-9]+)", r"Siege_Firepower[=：:]\s*([0-9]+)"],
        "command_value": [r"指挥值[=：:]\s*([0-9]+)", r"Command_Value[=：:]\s*([0-9]+)"],
        "service_limit": [r"服役上限[=：:]\s*([0-9]+)", r"Service_Limit[=：:]\s*([0-9]+)"],
        "length_m": [r"长度[=：:]\s*([0-9]+)\s*米", r"长度([0-9]+)米"],
        "storage": [r"仓储[=：:]\s*([0-9]+)"],
    }
    for key, pats in patterns.items():
        for p in pats:
            m = re.search(p, text, re.I)
            if m:
                kv[key] = m.group(1)
                break
    return kv


def main():
    index = load_csv(WIKI / "ships_index.csv")
    detail = load_csv(WIKI / "ships_detail.csv")
    fleet = load_csv(WIKI / "未央公约组织_fleet.csv")

    by_name: dict[str, dict] = {}

    for r in index:
        name = r.get("name") or ""
        if not name:
            continue
        by_name[name] = {
            "name": name,
            "star": r.get("star", ""),
            "ship_type": r.get("ship_type", ""),
            "faction": r.get("faction", ""),
            "weapon_type": r.get("weapon_type", ""),
            "attack_attr": r.get("attack_attr", ""),
            "position": r.get("position", ""),
            "acquisition": r.get("acquisition", ""),
            "source": "index",
        }

    for r in detail:
        name = r.get("name") or ""
        if not name:
            continue
        row = by_name.setdefault(name, {"name": name, "source": "detail"})
        for k in [
            "structure", "armor", "energy_res", "cruise_speed", "warp_speed",
            "metal", "crystal", "deuterium", "command_value", "service_limit",
            "ship_dps", "air_dps", "siege_dps", "length", "storage", "linked_terms", "summary",
        ]:
            if r.get(k):
                row[k] = r[k]
        # from idx_ fields
        for a, b in [
            ("idx_star", "star"),
            ("idx_ship_type", "ship_type"),
            ("idx_faction", "faction"),
            ("idx_weapon_type", "weapon_type"),
            ("idx_attack_attr", "attack_attr"),
            ("idx_position", "position"),
            ("idx_acquisition", "acquisition"),
        ]:
            if r.get(a) and not row.get(b):
                row[b] = r[a]
        # reparse summary/blob
        blob = " ".join(str(r.get(x) or "") for x in r.keys())
        parsed = parse_numbers_from_text(blob)
        for k, v in parsed.items():
            if not row.get(k):
                row[k] = v
        row["source"] = (row.get("source") or "") + "+detail"

    # fleet sample mapping
    field_map = {
        "Ship_Name": "name",
        "Star_Rating": "star",
        "Type": "ship_type",
        "Weapon_Type": "weapon_type",
        "Attack_Attribute": "attack_attr",
        "Key_Skills": "skills",
        "Acquisition": "acquisition",
        "Position": "position",
        "Structure_Value": "structure",
        "Resistance_or_Armor": "armor",
        "Shield": "shield",
        "Speed_Cruise": "cruise_speed",
        "Speed_Curvature": "warp_speed",
        "Build_Metal": "metal",
        "Build_Crystal": "crystal",
        "Build_Heavy_Hydrogen": "deuterium",
        "Firepower_vs_Ship": "ship_dps",
        "Firepower_vs_Air": "air_dps",
        "Siege_Firepower": "siege_dps",
        "Command_Value": "command_value",
        "Service_Limit": "service_limit",
        "Notes": "notes",
    }
    for r in fleet:
        name = r.get("Ship_Name") or r.get("name")
        if not name:
            continue
        row = by_name.setdefault(name, {"name": name, "source": "fleet"})
        for src, dst in field_map.items():
            if src == "Ship_Name":
                continue
            val = (r.get(src) or "").strip()
            if not val or val == "TBD":
                continue
            # fleet 手填优先覆盖空，或当更完整时覆盖
            if not row.get(dst) or (len(val) > len(str(row.get(dst) or "")) and any(ch.isdigit() for ch in val)):
                # 清洗括号说明，尽量保留主数字
                row[dst] = val
        # 再解析 notes
        parsed = parse_numbers_from_text(" ".join(str(v or "") for v in r.values()))
        for k, v in parsed.items():
            if not row.get(k):
                row[k] = v
        row["faction"] = row.get("faction") or "未央公约组织"
        row["source"] = (row.get("source") or "") + "+fleet"

    rows = sorted(by_name.values(), key=lambda x: (x.get("star") or "9", x.get("ship_type") or "", x.get("name") or ""))

    fields = [
        "name", "star", "ship_type", "faction", "weapon_type", "attack_attr", "position", "acquisition",
        "structure", "armor", "shield", "cruise_speed", "warp_speed",
        "metal", "crystal", "deuterium", "command_value", "service_limit",
        "ship_dps", "air_dps", "siege_dps", "length", "length_m", "storage",
        "skills", "notes", "linked_terms", "summary", "source",
    ]
    out_csv = WIKI / "ships_enriched.csv"
    with out_csv.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)
    print("wrote", out_csv, len(rows))

    # stats
    for k in ["structure", "metal", "crystal", "deuterium", "command_value", "ship_dps", "cruise_speed"]:
        n = sum(1 for r in rows if r.get(k))
        print(f"  filled {k}: {n}/{len(rows)}")

    # xlsx
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    )
    ws = wb.active
    ws.title = "舰船增强总表"
    for c, h in enumerate(fields, 1):
        cell = ws.cell(1, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
    for r_i, r in enumerate(rows, 2):
        for c_i, h in enumerate(fields, 1):
            cell = ws.cell(r_i, c_i, r.get(h, ""))
            cell.border = thin
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for c in range(1, len(fields) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.column_dimensions["A"].width = 30
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(fields))}{len(rows)+1}"

    ws2 = wb.create_sheet("覆盖率")
    ws2.append(["字段", "已填行数", "总行数", "覆盖率%"])
    for k in fields:
        n = sum(1 for r in rows if r.get(k))
        ws2.append([k, n, len(rows), round(100 * n / max(len(rows), 1), 1)])

    xlsx_path = DATA / "舰船数据总表.xlsx"
    wb.save(xlsx_path)
    print("wrote", xlsx_path)


if __name__ == "__main__":
    main()
