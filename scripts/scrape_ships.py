#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从 biligame Wiki 拉取舰船索引与尽可能多的详细属性。
输出：
  wiki_data/ships_index.csv
  wiki_data/ships_detail.csv
  data/舰船数据总表.xlsx
"""
from __future__ import annotations

import csv
import re
import time
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
WIKI_DIR = ROOT / "wiki_data"
DATA_DIR = ROOT / "data"
API = "https://wiki.biligame.com/wjdlglr/api.php"

session = requests.Session()
session.headers.update(
    {
        "User-Agent": "guoyunkaituo-kb/1.0 (novel writing research; contact: local; respects crawl delay)",
        "Accept-Language": "zh-CN,zh;q=0.9",
    }
)


def api_get(params: dict, retries=3):
    params = {**params, "format": "json"}
    last = None
    for i in range(retries):
        try:
            r = session.get(API, params=params, timeout=60)
            r.raise_for_status()
            # 有时返回 HTML 拦截
            if "application/json" not in r.headers.get("Content-Type", "") and not r.text.strip().startswith("{"):
                raise ValueError(f"non-json response: {r.text[:200]}")
            return r.json()
        except Exception as e:
            last = e
            time.sleep(1.5 * (i + 1))
    raise last


def list_category_members(title: str) -> list[str]:
    members = []
    cont = None
    while True:
        params = {
            "action": "query",
            "list": "categorymembers",
            "cmtitle": title,
            "cmlimit": "max",
            "cmnamespace": 0,
        }
        if cont:
            params["cmcontinue"] = cont
        data = api_get(params)
        members.extend([m["title"] for m in data.get("query", {}).get("categorymembers", [])])
        cont = data.get("continue", {}).get("cmcontinue")
        if not cont:
            break
        time.sleep(0.3)
    return members


def parse_ship_index_from_gallery() -> list[dict]:
    """解析舰船图鉴 HTML 表格，拿星级/武器/势力等索引字段。"""
    data = api_get({"action": "parse", "page": "舰船图鉴", "prop": "text"})
    html = data["parse"]["text"]["*"]
    soup = BeautifulSoup(html, "html.parser")
    rows_out = []
    # 找所有表格行
    for tr in soup.select("table tr"):
        tds = tr.find_all(["td", "th"])
        if len(tds) < 6:
            continue
        texts = [" ".join(td.stripped_strings) for td in tds]
        # 名称列通常含舰船链接
        link = tr.find("a", href=True)
        if not link:
            continue
        name = link.get("title") or link.get_text(strip=True)
        if not name or name in ("预览", "名称") or "分类" in name:
            continue
        # 启发式：星级字段
        star = ""
        for t in texts:
            m = re.search(r"([1-6])星", t)
            if m:
                star = m.group(1) + "星"
                break
        def find_one(cands):
            for t in texts:
                for c in cands:
                    if c == t or c in t.split():
                        return c
            for t in texts:
                for c in cands:
                    if c in t:
                        return c
            return ""

        weapon = find_one(["火炮", "轨道炮", "离子炮", "导弹", "鱼雷", "脉冲炮", "等离子武器", "飞行器", "无"])
        attr = find_one(["实弹", "能量", "未定", "无"])
        pos = find_one(["前排", "中排", "后排", "其他"])
        stype = find_one(
            ["工程舰", "战机", "护航艇", "护卫舰", "驱逐舰", "巡洋舰", "战列巡洋舰", "支援舰", "航空母舰", "战列舰", "登陆舰", "英雄舰", "民用舰船"]
        )
        faction = find_one(
            [
                "未央公约组织",
                "安东尼奥斯财团",
                "诺玛运输集团",
                "盘古拓展集团",
                "木星工业集团",
                "雷火科技公司",
                "海雷丁家族",
                "神圣群星帝国",
                "流放者组织",
                "其他",
            ]
        )
        acq = find_one(["常驻", "限定", "合成", "无"])
        rows_out.append(
            {
                "name": name,
                "star": star,
                "weapon_type": weapon,
                "attack_attr": attr,
                "position": pos,
                "ship_type": stype,
                "faction": faction,
                "acquisition": acq,
                "raw_cells": " | ".join(texts[:12]),
            }
        )
    # 去重保序
    seen = set()
    uniq = []
    for r in rows_out:
        if r["name"] in seen:
            continue
        seen.add(r["name"])
        uniq.append(r)
    return uniq


def extract_kv_from_wikitext(wt: str) -> dict:
    """从舰船页 wikitext 抽关键数值（启发式）。"""
    kv = {}
    patterns = {
        "structure": r"结构值\s*[：:|=]\s*([0-9,\.]+)",
        "armor": r"装甲\s*[：:|=]\s*([0-9,\.%]+)",
        "energy_res": r"能量抗性\s*[：:|=]\s*([0-9,\.%]+)",
        "cruise_speed": r"巡航速度\s*[：:|=]\s*([0-9,\.]+)",
        "warp_speed": r"曲率速度\s*[：:|=]\s*([0-9,\.]+)",
        "metal": r"金属\s*[：:|=]\s*([0-9,\.]+)",
        "crystal": r"晶体\s*[：:|=]\s*([0-9,\.]+)",
        "deuterium": r"重氢\s*[：:|=]\s*([0-9,\.]+)",
        "command_value": r"指挥值\s*[：:|=]\s*([0-9,\.]+)",
        "service_limit": r"(?:服役|生产)上限\s*[：:|=]\s*([0-9,\.]+)",
        "ship_dps": r"对舰\s*[：:|=]\s*([0-9,\.]+)",
        "air_dps": r"对空\s*[：:|=]\s*([0-9,\.]+)",
        "siege_dps": r"攻城\s*[：:|=]\s*([0-9,\.]+)",
        "length": r"长度\s*[：:|=]\s*([0-9,\.]+)\s*米?",
        "storage": r"仓储\s*[：:|=]\s*([0-9,\.]+)",
    }
    for k, p in patterns.items():
        m = re.search(p, wt)
        if m:
            kv[k] = m.group(1).replace(",", "")

    # 模板参数风格 |结构值=123
    for k, label in [
        ("structure", "结构值"),
        ("armor", "装甲"),
        ("cruise_speed", "巡航速度"),
        ("warp_speed", "曲率速度"),
        ("metal", "金属"),
        ("crystal", "晶体"),
        ("deuterium", "重氢"),
        ("command_value", "指挥值"),
    ]:
        if k not in kv:
            m = re.search(rf"\|\s*{label}\s*=\s*([^\|\n]+)", wt)
            if m:
                kv[k] = m.group(1).strip()

    # 技能/策略：收集看起来像技能的列表
    skills = re.findall(r"\[\[([^\[\]]+?)\]\]", wt)
    # 过滤噪声
    noise = {"首页", "舰船图鉴", "分类"}
    skills = [s.split("|")[-1] for s in skills if s not in noise and len(s) < 30][:20]
    if skills:
        kv["linked_terms"] = "、".join(skills[:15])
    return kv


def fetch_ship_detail(title: str) -> dict:
    data = api_get({"action": "parse", "page": title, "prop": "wikitext"})
    wt = data.get("parse", {}).get("wikitext", {}).get("*", "")
    kv = extract_kv_from_wikitext(wt)
    kv["name"] = title
    kv["wikitext_len"] = len(wt)
    # 截断摘要
    plain = re.sub(r"<[^>]+>", "", wt)
    plain = re.sub(r"\{\{[^\}]+\}\}", " ", plain)
    plain = re.sub(r"\[\[([^|\]]+\|)?([^\]]+)\]\]", r"\2", plain)
    plain = re.sub(r"\s+", " ", plain).strip()
    kv["summary"] = plain[:300]
    return kv


def write_csv(path: Path, rows: list[dict], fieldnames: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)
    print("wrote", path, "rows", len(rows))


def to_xlsx(index_rows, detail_rows, path: Path):
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def dump(ws, rows, fields):
        for c, h in enumerate(fields, 1):
            cell = ws.cell(1, c, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin
        for r_i, row in enumerate(rows, 2):
            for c_i, h in enumerate(fields, 1):
                cell = ws.cell(r_i, c_i, row.get(h, ""))
                cell.border = thin
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for c in range(1, len(fields) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 14
        ws.freeze_panes = "A2"
        if rows:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(fields))}{len(rows)+1}"

    ws1 = wb.active
    ws1.title = "舰船索引"
    idx_fields = ["name", "star", "ship_type", "faction", "weapon_type", "attack_attr", "position", "acquisition", "raw_cells"]
    dump(ws1, index_rows, idx_fields)
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["I"].width = 40

    ws2 = wb.create_sheet("详细属性")
    det_fields = [
        "name", "structure", "armor", "energy_res", "cruise_speed", "warp_speed",
        "metal", "crystal", "deuterium", "command_value", "service_limit",
        "ship_dps", "air_dps", "siege_dps", "length", "storage", "linked_terms", "summary",
    ]
    dump(ws2, detail_rows, det_fields)
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["Q"].width = 30
    ws2.column_dimensions["R"].width = 50

    ws3 = wb.create_sheet("字段说明")
    notes = [
        ["字段", "含义", "备注"],
        ["structure", "结构值", "近似HP"],
        ["armor", "装甲/物理抗性", "Wiki表述不一"],
        ["cruise_speed / warp_speed", "巡航/曲率速度", ""],
        ["metal/crystal/deuterium", "建造资源", "可能含强化后数值"],
        ["command_value", "指挥值", "编队容量占用"],
        ["ship_dps/air_dps/siege_dps", "对舰/对空/攻城", "面板火力参考"],
        ["来源", "BWiki 舰船图鉴及分页面", "CC BY-NC-SA 3.0；仅供创作参考"],
        ["缺失值", "空单元格", "页面结构差异或未实装字段；写作可用等级描述代替"],
    ]
    for r in notes:
        ws3.append(r)
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 40

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print("wrote", path)


def main():
    WIKI_DIR.mkdir(parents=True, exist_ok=True)
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    print("Listing category 舰船 ...")
    try:
        cat_ships = list_category_members("分类:舰船")
    except Exception as e:
        print("category list failed:", e)
        cat_ships = []
    print("category ships:", len(cat_ships))

    print("Parsing 舰船图鉴 index ...")
    try:
        index_rows = parse_ship_index_from_gallery()
    except Exception as e:
        print("gallery parse failed:", e)
        index_rows = []
    print("index rows:", len(index_rows))

    # 合并分类中未出现在索引的名字
    idx_names = {r["name"] for r in index_rows}
    for name in cat_ships:
        if name not in idx_names:
            index_rows.append(
                {
                    "name": name,
                    "star": "",
                    "weapon_type": "",
                    "attack_attr": "",
                    "position": "",
                    "ship_type": "",
                    "faction": "",
                    "acquisition": "",
                    "raw_cells": "from_category_only",
                }
            )
            idx_names.add(name)

    write_csv(
        WIKI_DIR / "ships_index.csv",
        index_rows,
        ["name", "star", "ship_type", "faction", "weapon_type", "attack_attr", "position", "acquisition", "raw_cells"],
    )

    # 详细页：优先拉有索引的，最多 N 艘避免过久；默认全量但限速
    detail_rows = []
    targets = [r["name"] for r in index_rows]
    # 也合并本地已有 fleet 样本名
    local_fleet = WIKI_DIR / "未央公约组织_fleet.csv"
    if local_fleet.exists():
        with local_fleet.open(encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                n = row.get("Ship_Name") or row.get("name")
                if n and n not in targets:
                    targets.append(n)

    print(f"Fetching details for {len(targets)} ships (rate-limited)...")
    for i, title in enumerate(targets, 1):
        try:
            d = fetch_ship_detail(title)
            # 合并索引信息
            for r in index_rows:
                if r["name"] == title:
                    d.update({f"idx_{k}": v for k, v in r.items() if k != "name"})
                    break
            detail_rows.append(d)
        except Exception as e:
            detail_rows.append({"name": title, "summary": f"ERROR: {e}"})
        if i % 20 == 0:
            print(f"  {i}/{len(targets)}")
        time.sleep(0.35)

    write_csv(
        WIKI_DIR / "ships_detail.csv",
        detail_rows,
        [
            "name", "structure", "armor", "energy_res", "cruise_speed", "warp_speed",
            "metal", "crystal", "deuterium", "command_value", "service_limit",
            "ship_dps", "air_dps", "siege_dps", "length", "storage", "linked_terms", "summary",
            "idx_star", "idx_ship_type", "idx_faction", "idx_weapon_type", "idx_attack_attr",
            "idx_position", "idx_acquisition",
        ],
    )
    to_xlsx(index_rows, detail_rows, DATA_DIR / "舰船数据总表.xlsx")

    # 写一个 markdown 索引方便 Obsidian
    md_path = ROOT / "obsidian_vault" / "04-舰船数据" / "舰船索引列表.md"
    md_path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "---",
        "tags: [舰船, 索引]",
        "---",
        "",
        f"# 舰船索引列表（共 {len(index_rows)}）",
        "",
        "数据生成自 Wiki API，详见 `wiki_data/ships_index.csv`。",
        "",
        "| 名称 | 星级 | 类型 | 势力 | 武器 | 获取 |",
        "|---|---|---|---|---|---|",
    ]
    for r in index_rows:
        lines.append(
            f"| {r.get('name','')} | {r.get('star','')} | {r.get('ship_type','')} | {r.get('faction','')} | {r.get('weapon_type','')} | {r.get('acquisition','')} |"
        )
    md_path.write_text("\n".join(lines), encoding="utf-8")
    print("wrote", md_path)
    print("SCRAPE DONE")


if __name__ == "__main__":
    main()
